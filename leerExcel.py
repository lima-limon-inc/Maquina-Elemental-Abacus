from openpyxl import Workbook, load_workbook

def leerAbacusExcel(archivo, dondeComienza=1):
    wb = load_workbook(archivo, data_only=True).active
    with open((str(archivo.split(".",1)[0]) + ".abs"), "w") as f: #Odio que haya tanta indentacion pero es la unica forma que se me ocurre de hacerlo para que sea modular y no tengas que descargar la libreria si no la usas. Esto crea un nuevo archivo que el excel pasada, solo que con terminacion abs
        for fila in wb.iter_rows():
            textoFila = ""
            # print(textoFila)
            for i in range(0,wb.max_column): #TODO: Hacer constante
                valorCelda = fila[i].value

                if valorCelda == None or valorCelda == " ":
                    continue

                textoFila += str(valorCelda) + ","
            f.write(textoFila + "\n")
# leerAbacusExcel()


